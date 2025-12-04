import streamlit as st
import base64 
from datetime import datetime
import os, tempfile, webbrowser
import pandas as pd
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage

# ---------- CONFIG ----------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
IMAGE_DIR = os.path.join(BASE_DIR, "images")  # folder to store all reference images

EXCEL = os.path.join(BASE_DIR, "test6.xlsx")        # <-- final workbook name
SHEET_MB = "Mixing Block Data"
SHEET_GW = "Gas-Water Block Data"
SHEET_SPECS = "Specs"
SHEET_GW_REF = "GW Reference Photo"
SHEET_MI_REF = "MI Reference Photo"
GW_IMAGE = os.path.join(IMAGE_DIR, "TestGWB.png")    # <-- reference image filename

PART_SHORT = {"mi": "Mixing Block", "gw": "Gas/Water Block"}
MACHINE_MAP = {"1": "SA001", "2": "SA002", "3": "SA003"}
TOL_DEFAULT = 0.50

SPECS = {
    "Mixing Block": {
        "1": {"Inner": (4.00, 3.50, 4.50), "Outer": (9.00, 8.50, 9.50)},
        "2": {"Inner": (4.00, 3.50, 4.50), "Outer": (9.00, 8.50, 9.50)},
        "3": {"Inner": (6.40, 5.90, 6.90), "Outer": (9.20, 8.70, 9.70)},
        "4": {"Inner": (9.20, 8.70, 9.70), "Outer": (12.80, 12.30, 13.30)},
    },
    "Gas/Water Block": {
        "1": {"Inner": (6.00, 5.50, 6.50)},
        "2": {"Inner": (6.00, 5.50, 6.50)},
        "3": {"Inner": (6.00, 5.50, 6.50)},
        "4": {"Inner": (6.00, 5.50, 6.50)},
        "5": {"Inner": (6.00, 5.50, 6.50)},
    },
}

DATA_COLS = [
    "Timestamp","Measured Date", "Machine", "Part Type", "Chamber", "Piece ID", "Part In/Out", "Batch Cleaning",
    "Hole", "Feature", "Value",
    "Nominal", "LSL", "USL", "Status", "Notes"
]

SPECS_COLS = ["Part Type", "Hole", "Feature", "Nominal", "LSL", "USL", "Tolerance"]

# ---------- Utilities ----------
def atomic_write_all(filename, sheets_dict):
    fd, tmp = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    try:
        with pd.ExcelWriter(tmp, engine="openpyxl") as w:
            for sheet_name, df in sheets_dict.items():
                df.to_excel(w, sheet_name=sheet_name, index=False)
        os.replace(tmp, filename)
        return filename, None
    except PermissionError:
        try: os.remove(tmp)
        except: pass
        alt = f"{os.path.splitext(filename)[0]}_LOCKED_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        try:
            with pd.ExcelWriter(alt, engine="openpyxl") as w:
                for sheet_name, df in sheets_dict.items():
                    df.to_excel(w, sheet_name=sheet_name, index=False)
            return None, alt
        except Exception:
            try: os.remove(alt)
            except: pass
            return None, None
    except Exception:
        try: os.remove(tmp)
        except: pass
        return None, None

def read_sheet_safe(sheet_name):
    if not os.path.exists(EXCEL):
        return pd.DataFrame(columns=DATA_COLS if sheet_name != SHEET_SPECS else SPECS_COLS)
    try:
        return pd.read_excel(EXCEL, sheet_name=sheet_name)
    except Exception:
        return pd.DataFrame(columns=DATA_COLS if sheet_name != SHEET_SPECS else SPECS_COLS)

def build_specs_df():
    rows = []
    for part, holes in SPECS.items():
        for hole, feats in holes.items():
            for feat, (nom, lsl, usl) in feats.items():
                rows.append({
                    "Part Type": part,
                    "Hole": f"H{hole}",
                    "Feature": feat,
                    "Nominal": nom,
                    "LSL": lsl,
                    "USL": usl,
                    "Tolerance": round(nom - lsl, 4)
                })
    return pd.DataFrame(rows, columns=SPECS_COLS)

def ensure_workbook():
    if not os.path.exists(EXCEL):
        sheets = {
            SHEET_MB: pd.DataFrame(columns=DATA_COLS),
            SHEET_GW: pd.DataFrame(columns=DATA_COLS),
            SHEET_SPECS: build_specs_df(),
            SHEET_GW_REF: pd.DataFrame(["Image Loads Below"])
        }
        saved, alt = atomic_write_all(EXCEL, sheets)
        if saved:
            try: add_reference_image()
            except: pass
        return

    try:
        all_sheets = pd.read_excel(EXCEL, sheet_name=None)
    except Exception:
        try: os.remove(EXCEL)
        except: pass
        return ensure_workbook()

    changed = False
    if SHEET_MB not in all_sheets:
        all_sheets[SHEET_MB] = pd.DataFrame(columns=DATA_COLS); changed = True
    if SHEET_GW not in all_sheets:
        all_sheets[SHEET_GW] = pd.DataFrame(columns=DATA_COLS); changed = True
    if SHEET_GW_REF not in all_sheets:
        all_sheets[SHEET_GW_REF] = pd.DataFrame(["Image Loads Below"]); changed = True

    all_sheets[SHEET_SPECS] = build_specs_df()
    changed = True

    if changed:
        saved, alt = atomic_write_all(EXCEL, all_sheets)
        if saved:
            try: add_reference_image()
            except: pass
            try: apply_excel_coloring_and_separator([SHEET_MB, SHEET_GW])
            except: pass

# ---------- NEW: helper for Streamlit UI ----------
def get_specs_df(part_filter=None):
    ensure_workbook()
    df = read_sheet_safe(SHEET_SPECS)
    
    if part_filter and part_filter != "All":
        df = df[df["Part Type"] == part_filter]

    return df

# ✅---------- FIXED export for vendor ----------
def export_specs_for_vendor(filename="specs.csv", part_filter=None):
    ensure_workbook()
    df = get_specs_df(part_filter)

    if df.empty:
        return False, "No specs available for export"

    for c in ["Nominal", "LSL", "USL", "Tolerance"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")

    df.to_csv(filename, index=False)
    return True, os.path.abspath(filename)

# ---------- Excel formatting ----------
def apply_excel_coloring_and_separator(data_sheets):
    if not os.path.exists(EXCEL):
        return
    try:
        wb = load_workbook(EXCEL)
    except Exception:
        return

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thick_side = Side(border_style="medium", color="000000")
    def set_bottom_border(ws, row_idx, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row_idx, column=c)
            top = cell.border.top; left = cell.border.left; right = cell.border.right
            cell.border = Border(top=top, left=left, right=right, bottom=thick_side)

    for sheet_name in data_sheets:
        if sheet_name not in wb.sheetnames: continue
        ws = wb[sheet_name]
        try:
            header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        except StopIteration:
            continue
        ncols = max(1, len(header))
        try:
            status_col_idx = header.index("Status") + 1
        except ValueError:
            status_col_idx = None
        timestamp_idx = header.index("Timestamp") + 1 if "Timestamp" in header else None
        pieceid_idx = header.index("Piece ID") + 1 if "Piece ID" in header else None

        prev_group = None
        for r in range(2, ws.max_row + 1):
            if status_col_idx:
                status_cell = ws.cell(row=r, column=status_col_idx)
                if status_cell.value and str(status_cell.value).strip().upper() == "FAIL":
                    for c in range(1, ncols + 1):
                        ws.cell(row=r, column=c).fill = red_fill
            if timestamp_idx and pieceid_idx:
                ts_val = ws.cell(row=r, column=timestamp_idx).value
                pid_val = ws.cell(row=r, column=pieceid_idx).value
                group_key = (ts_val, pid_val)
            else:
                pid_val = ws.cell(row=r, column=pieceid_idx).value if pieceid_idx else None
                group_key = (None, pid_val)
            if prev_group is not None and group_key != prev_group:
                set_bottom_border(ws, r - 1, ncols)
            prev_group = group_key
        if ws.max_row >= 2:
            set_bottom_border(ws, ws.max_row, ncols)

    try: wb.save(EXCEL)
    except: pass
    wb.close()

from openpyxl.utils import get_column_letter

from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import os

from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import os

from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import os

def add_reference_image():
    if not os.path.exists(EXCEL):
        return
    try:
        wb = load_workbook(EXCEL)
    except:
        return

    # --- GW Reference ---
    gw_sheet_name = "GW Reference Photo"
    ws_gw = wb[gw_sheet_name] if gw_sheet_name in wb.sheetnames else wb.create_sheet(gw_sheet_name)
    ws_gw.sheet_view.showGridLines = False

    gw_imgs = [os.path.join(IMAGE_DIR, "TestGWB.png"), os.path.join(IMAGE_DIR, "TestGWBtri.png")]
    row_start = 1
    for img_path in gw_imgs:
        if os.path.exists(img_path):
            img = XLImage(img_path)
            ws_gw.add_image(img, f"A{row_start}")
            ws_gw.row_dimensions[row_start].height = img.height * 0.75
            row_start += max(int(img.height / 15), 1)

    # --- MI Reference ---
    mi_sheet_name = "MI Reference Photo"
    ws_mi = wb[mi_sheet_name] if mi_sheet_name in wb.sheetnames else wb.create_sheet(mi_sheet_name)
    ws_mi.sheet_view.showGridLines = False

    mi_images = [os.path.join(IMAGE_DIR, "MBtop.png"), os.path.join(IMAGE_DIR, "MBbot.png"), os.path.join(IMAGE_DIR, "MBtri.png")]
    col_start = 1
    for img_file in mi_images:
        if os.path.exists(img_file):
            img = XLImage(img_file)
            col_letter = get_column_letter(col_start)
            ws_mi.add_image(img, f"{col_letter}1")
            ws_mi.column_dimensions[col_letter].width = img.width / 7
            ws_mi.row_dimensions[1].height = max(ws_mi.row_dimensions[1].height or 15, img.height * 0.75)
            col_start += max(int(img.width / 60), 3)

    try:
        wb.save(EXCEL)
    except:
        pass
    wb.close()

# ---------- Core API for Streamlit ----------
def _status_from_value(part, hole, feat, val):
    try:
        nominal, lsl, usl = SPECS[part][hole][feat]
        return "PASS" if (lsl <= val <= usl) else "FAIL", nominal, lsl, usl
    except Exception:
        return ("PASS" if val is not None else "FAIL", None, None, None)

def add_measurement_rows(part, machine, chamber, piece_id, part_flow, notes, measurements, measured_date=None, batch_number=None, timestamp=None):
    import streamlit as st  # ensure Streamlit is available

    ensure_workbook()
    ts = timestamp if timestamp is not None else datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ---------------------- USE SESSION CACHE ----------------------
    if "excel_cache" not in st.session_state:
        st.session_state["excel_cache"] = {}

    cache = st.session_state["excel_cache"]

    # Load part sheet from cache or Excel
    sheet_name = SHEET_MB if part == "Mixing Block" else SHEET_GW
    other_sheet_name = SHEET_GW if part == "Mixing Block" else SHEET_MB

    if sheet_name in cache:
        df_part = cache[sheet_name]
    else:
        df_part = read_sheet_safe(sheet_name)
        cache[sheet_name] = df_part

    if other_sheet_name in cache:
        df_other = cache[other_sheet_name]
    else:
        df_other = read_sheet_safe(other_sheet_name)
        cache[other_sheet_name] = df_other

    if SHEET_SPECS in cache:
        df_specs = cache[SHEET_SPECS]
    else:
        df_specs = read_sheet_safe(SHEET_SPECS)
        cache[SHEET_SPECS] = df_specs
    # ---------------------------------------------------------------

    # Format measured date to only YYYY-MM-DD
    if measured_date is None:
        measured_date_str = datetime.now().strftime("%Y-%m-%d")
    else:
        measured_date_str = str(measured_date)
        if isinstance(measured_date, datetime):
            measured_date_str = measured_date.strftime("%Y-%m-%d")

    rows = []
    for m in measurements:
        hole = str(m.get("Hole")).lstrip("H")
        feat = m.get("Feature")
        try:
            val = float(m.get("Value"))
        except:
            val = None
        status, nominal, lsl, usl = _status_from_value(part, hole, feat, val if val is not None else 0.0)

        # ⭐ Add image path if exists in measurement dict
        img_path = m.get("ImagePath", None)

        rows.append({
            "Timestamp": ts,
            "Measured Date": measured_date_str,
            "Machine": machine,
            "Part Type": part,
            "Part In/Out": part_flow,
            "Batch Cleaning": batch_number if batch_number is not None else "",
            "Chamber": chamber,
            "Piece ID": piece_id,
            "Hole": f"H{hole}",
            "Feature": feat,
            "Value": val,
            "Nominal": nominal,
            "LSL": lsl,
            "USL": usl,
            "Status": status,
            "Notes": notes,
            "Image Path": img_path
        })

    if not rows:
        return False, "No rows to add"

    df_append = pd.DataFrame(rows)
    for c in ["Value", "Nominal", "LSL", "USL"]:
        if c in df_append.columns:
            df_append[c] = pd.to_numeric(df_append[c], errors="coerce")

    df_part = pd.concat([df_part, df_append], ignore_index=True)

    # ---------------------- UPDATE CACHE BEFORE WRITING ----------------------
    cache[sheet_name] = df_part

    sheets = {
        SHEET_MB: cache[SHEET_MB],
        SHEET_GW: cache[SHEET_GW],
        SHEET_SPECS: cache[SHEET_SPECS]
    }
    # ------------------------------------------------------------------------

    saved, alt = atomic_write_all(EXCEL, sheets)
    if saved:
        try:
            add_reference_image()
        except:
            pass
        try:
            apply_excel_coloring_and_separator([SHEET_MB, SHEET_GW])
        except:
            pass
        return True, saved
    elif alt:
        return False, f"Excel locked — saved clone to: {alt}"
    else:
        return False, "Failed to save measurements"

    
def get_available_holes_for_part(part):
    """Return hole list depending on part type."""
    if part.lower() == "mixing block":
        return ["1", "2", "3", "4"]
    elif part.lower() in ["gas/water block", "gas-water block", "gas water block"]:
        return ["1", "2", "3", "4", "5"]
    else:
        return ["1"]


def get_features_for_part(part, hole):
    if part == "Mixing Block":
        return ["Inner", "Outer"]
    else:
        return ["Inner"]

# ---------------- SAFE HOLE SORT KEY ----------------
def safe_hole_sort_key(x):
    """
    Convert a hole identifier into an integer sort key.

    Accepts values like:
      -  "1", "2", ...
      -  "H1", "H2"
      -  "H10"
      -  "meas_H1_Inner", "meas_H10_Outer", "someprefix_H3_suffix"
      -  other noisy strings containing a number

    Returns:
      - integer hole number when found (so sorting is numeric)
      - 999 as a safe fallback for unknown values
    """
    try:
        s = str(x).strip()
    except Exception:
        return 999

    # Quick pure-digit
    if s.isdigit():
        return int(s)

    # H-prefixed like "H1" or "h2"
    if len(s) >= 2 and (s[0].upper() == "H") and s[1:].isdigit():
        return int(s[1:])

    # Find the first integer anywhere in the string (e.g. "meas_H12_Inner")
    import re
    m = re.search(r'(\d+)', s)
    if m:
        try:
            return int(m.group(1))
        except Exception:
            return 999

    # fallback (unknown)
    return 999


def show_trend_df(part, machine=None, chamber=None, hole=None, feature=None):
    df = read_sheet_safe(SHEET_MB if part == "Mixing Block" else SHEET_GW)
    if df is None or df.empty:
        return pd.DataFrame()
    if machine:
        df = df[df["Machine"] == machine]
    if chamber:
        df = df[df["Chamber"] == chamber]
    if hole:
        df = df[df["Hole"] == hole]
    if feature:
        df = df[df["Feature"] == feature]
    if df.empty:
        return pd.DataFrame()
    try: df["Timestamp"] = pd.to_datetime(df["Timestamp"])
    except: pass
    return df.sort_values("Timestamp")

def delete_rows_by_indexes(part, indexes_str):
    """
    Final fixed version with session cache:
    ✅ Handles 0 correctly
    ✅ Deletes last row in any range
    ✅ Works with ranges like 0-5, 1-3, 0, 3,5-7
    ✅ Safe Excel rewrite
    ✅ Uses st.rerun() (no experimental_rerun)
    ✅ Prevents data loss when switching holes
    """

    import streamlit as st

    # ---------------------- USE SESSION CACHE ----------------------
    if "excel_cache" not in st.session_state:
        st.session_state["excel_cache"] = {}
    cache = st.session_state["excel_cache"]

    sheet_name = SHEET_MB if part == "Mixing Block" else SHEET_GW
    other_sheet_name = SHEET_GW if part == "Mixing Block" else SHEET_MB

    # Load sheets from cache or Excel
    if sheet_name in cache:
        df = cache[sheet_name]
    else:
        df = read_sheet_safe(sheet_name)
        cache[sheet_name] = df

    if other_sheet_name not in cache:
        df_other = read_sheet_safe(other_sheet_name)
        cache[other_sheet_name] = df_other
    else:
        df_other = cache[other_sheet_name]

    if SHEET_SPECS not in cache:
        specs_df = read_sheet_safe(SHEET_SPECS)
        cache[SHEET_SPECS] = specs_df
    else:
        specs_df = cache[SHEET_SPECS]
    # ---------------------------------------------------------------

    if df.empty:
        return False, "No data found in sheet."

    # --- Parse indexes from string ---
    raw_input = str(indexes_str).strip()
    if not raw_input:
        return False, "No indexes entered."

    indexes = set()
    for part_str in raw_input.split(","):
        part_str = part_str.strip()
        if not part_str:
            continue
        if "-" in part_str:
            try:
                start, end = map(int, part_str.split("-"))
                if start > end:
                    start, end = end, start
                indexes.update(range(start, end + 1))
            except ValueError:
                continue
        else:
            try:
                indexes.add(int(part_str))
            except ValueError:
                continue

    if not indexes:
        return False, "No valid indexes entered."

    indexes = sorted(list(indexes))
    n = len(df)

    # --- Validate index bounds ---
    valid_indexes = [i for i in indexes if 0 <= i < n]
    if not valid_indexes:
        return False, f"No valid rows found. Valid range: 0–{n-1}"

    # --- Drop rows correctly ---
    try:
        df_after = df.drop(index=valid_indexes, errors="ignore").reset_index(drop=True)
    except Exception as e:
        return False, f"Error deleting rows: {e}"

    # ---------------------- UPDATE CACHE BEFORE WRITING ----------------------
    cache[sheet_name] = df_after

    sheets = {
        SHEET_MB: cache[SHEET_MB],
        SHEET_GW: cache[SHEET_GW],
        SHEET_SPECS: cache[SHEET_SPECS]
    }
    # ------------------------------------------------------------------------

    # --- Save back to Excel ---
    try:
        saved, alt = atomic_write_all(EXCEL, sheets)
        if saved:
            try:
                apply_excel_coloring_and_separator([SHEET_MB, SHEET_GW])
            except Exception:
                pass

            st.session_state["last_delete_success"] = True
            st.session_state["last_delete_msg"] = f"✅ Deleted rows: {valid_indexes}"
            st.success(f"✅ Deleted rows: {valid_indexes}")

            # --- ✅ Modern Streamlit refresh ---
            import time
            time.sleep(0.5)
            try:
                st.rerun()
            except Exception:
                st.warning("✅ Deleted successfully. Please refresh manually if table not updated.")

            return True, f"✅ Deleted rows: {valid_indexes}"

        elif alt:
            return False, f"Excel locked — saved copy created: {alt}"
        else:
            return False, "Failed to save Excel changes."

    except Exception as e:
        return False, f"Unexpected error: {e}"



def open_excel_file():
    if not os.path.exists(EXCEL):
        return False, "Workbook not created yet"
    try:
        os.startfile(os.path.abspath(EXCEL))
        return True, "Opened with default app"
    except Exception:
        try:
            webbrowser.open(os.path.abspath(EXCEL))
            return True, "Opened via webbrowser"
        except Exception as e:
            return False, f"Cannot open: {e}"

def get_reference_image_path():
    if os.path.exists(GW_IMAGE):
        return os.path.abspath(GW_IMAGE)
    return None

def show_reference_photos():
    import streamlit as st
    import os, base64, webbrowser
    from PIL import Image

    st.subheader("📸 Reference Photos Viewer")

    # --- Use correct IMAGE_DIR from main code ---
    IMAGE_DIR = os.path.join(BASE_DIR, "images")
    PLACEHOLDER = os.path.join(IMAGE_DIR, "placeholder.png")

    # --- Step 1: Choose Part Type ---
    part_type = st.radio(
        "Select Part Type:",
        ["Mixing Block", "Gas/Water Block"],
        horizontal=True
    )

    # --- Step 2: Define valid image mappings ---
    image_map = {
        "Mixing Block": {
            "Top View": os.path.join(IMAGE_DIR, "MBtop.png"),
            "Bottom View": os.path.join(IMAGE_DIR, "MBbot.png"),
            "Trimetric View": os.path.join(IMAGE_DIR, "MBtri.png"),
        },
        "Gas/Water Block": {
            "Front View": os.path.join(IMAGE_DIR, "TestGWB.png"),
            "Trimetric View": os.path.join(IMAGE_DIR, "TestGWBtri.png"),
        },
    }
    
    # --- Step 3: Define view options ---
    view_options = list(image_map[part_type].keys()) + ["All Views (Side by Side)"]
    selected_view = st.selectbox("Select View:", view_options)

    st.markdown("<hr>", unsafe_allow_html=True)

    # --- Step 4: Display Logic ---
    def load_image_or_placeholder(path, placeholder=PLACEHOLDER):
        if os.path.exists(path):
            return Image.open(path)
        elif os.path.exists(placeholder):
            return Image.open(placeholder)
        else:
            return None  # no image available at all

    if selected_view == "All Views (Side by Side)":
        st.markdown(f"<h4 style='text-align:center;'>{part_type} — All Views</h4>", unsafe_allow_html=True)
        views = list(image_map[part_type].keys())
        num_cols = min(3, len(views))
        cols = st.columns(num_cols, gap="medium")

        # Determine resize sizes
        img_sizes = {}
        if part_type == "Mixing Block":
            max_width = 350
            for view in views:
                img = load_image_or_placeholder(image_map[part_type][view])
                if img:
                    ratio = max_width / img.width
                    img_sizes[view] = (max_width, int(img.height * ratio))
        else:
            target_height = 500
            for view in views:
                img = load_image_or_placeholder(image_map[part_type][view])
                if img:
                    ratio = target_height / img.height
                    img_sizes[view] = (int(img.width * ratio), target_height)

        # Display images
        for i, view in enumerate(views):
            img_path = image_map[part_type][view]
            img = load_image_or_placeholder(img_path)
            if img:
                w, h = img_sizes.get(view, (300, 300))
                img = img.resize((w, h))
                img_bytes = open(img_path, "rb").read() if os.path.exists(img_path) else open(PLACEHOLDER, "rb").read()
                img_base64 = base64.b64encode(img_bytes).decode()
                cols[i % num_cols].markdown(
                    f"""
                    <div style="text-align:center;">
                        <img src="data:image/png;base64,{img_base64}" 
                             alt="{view}" 
                             style="width:{w}px; height:{h}px; border-radius:10px; 
                             box-shadow:0px 4px 10px rgba(0,0,0,0.3);" />
                        <p style="font-style:italic; color:gray;">{view}</p>
                    </div>
                    """,
                    unsafe_allow_html=True
                )

    else:
        # Single Image View
        img_path = image_map[part_type][selected_view]
        img = load_image_or_placeholder(img_path)
        if img:
            if part_type == "Mixing Block":
                max_width = 600
                ratio = max_width / img.width
                new_height = int(img.height * ratio)
            else:
                new_height = 500
                ratio = new_height / img.height
                max_width = int(img.width * ratio)
            img = img.resize((max_width, new_height))
            img_bytes = open(img_path, "rb").read() if os.path.exists(img_path) else open(PLACEHOLDER, "rb").read()
            img_base64 = base64.b64encode(img_bytes).decode()
            st.markdown(
                f"""
                <div style="text-align:center;">
                    <img src="data:image/png;base64,{img_base64}" 
                         style="width:{max_width}px; height:{new_height}px; border-radius:10px; 
                         box-shadow:0px 4px 10px rgba(0,0,0,0.3);" />
                    <p style="font-style:italic; color:gray;">{part_type} — {selected_view}</p>
                </div>
                """,
                unsafe_allow_html=True
            )
        else:
            st.info(f"No image available for {selected_view}.")

    st.markdown("<hr>", unsafe_allow_html=True)

    # --- Step 5: Open Externally Section ---
    st.markdown("### 🔍 Open Externally")
    open_view = st.selectbox("Select View to Open:", list(image_map[part_type].keys()), key=f"open_{part_type}")
    if st.button("Open Selected Image", key=f"open_btn_{part_type}"):
        img_path = image_map[part_type][open_view]
        if os.path.exists(img_path):
            try:
                os.startfile(img_path)
            except Exception:
                webbrowser.open(img_path)
        elif os.path.exists(PLACEHOLDER):
            st.info(f"{open_view} image missing. Opening placeholder.")
            webbrowser.open(PLACEHOLDER)
        else:
            st.warning(f"⚠️ {open_view} image not found.")

    # --- Step 6: Download Section ---
    st.markdown("### ⬇️ Download Image")
    dl_view = st.selectbox("Select View to Download:", list(image_map[part_type].keys()), key=f"dl_{part_type}")
    img_file = image_map[part_type][dl_view]
    if os.path.exists(img_file):
        with open(img_file, "rb") as f:
            st.download_button(
                label=f"Download {dl_view}",
                data=f,
                file_name=os.path.basename(img_file),
                mime="image/png",
                key=f"dl_btn_{part_type}"
            )
    elif os.path.exists(PLACEHOLDER):
        with open(PLACEHOLDER, "rb") as f:
            st.download_button(
                label=f"Download {dl_view} (placeholder)",
                data=f,
                file_name=os.path.basename(PLACEHOLDER),
                mime="image/png",
                key=f"dl_btn_placeholder_{part_type}"
            )
    else:
        st.warning(f"⚠️ {dl_view} image not available for download.")


def draw_trend_with_spec(df, spec_min, spec_max, title="Trend Chart"):
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots()

    x = list(range(len(df)))
    y = df["Value"].tolist()

    # Color points based on spec
    colors = []
    for v in y:
        if spec_min is not None and v < spec_min:
            colors.append("red")   # below spec
        elif spec_max is not None and v > spec_max:
            colors.append("red")   # above spec
        else:
            colors.append("blue")  # in spec

    ax.plot(x, y, marker="o")
    ax.scatter(x, y, c=colors, s=80)

    # Draw spec lines
    if spec_min is not None:
        ax.axhline(spec_min, linestyle="--", label="Min Spec", color="gray")
    if spec_max is not None:
        ax.axhline(spec_max, linestyle="--", label="Max Spec", color="gray")

    ax.set_title(title)
    ax.set_xlabel("Record #")
    ax.set_ylabel("Measurement (mm)")
    ax.legend()

    st.pyplot(fig)

    # Build analysis text
    issues = []
    for idx, v in enumerate(y):
        if spec_min is not None and v < spec_min:
            issues.append(f"• Point {idx+1}: **{v} mm** (below MIN spec {spec_min})")
        if spec_max is not None and v > spec_max:
            issues.append(f"• Point {idx+1}: **{v} mm** (above MAX spec {spec_max})")

    if issues:
        st.error("📛 **Out-of-Spec Points Detected**")
        for i in issues:
            st.write(i)
    else:
        st.success("✅ All points are within the specification.")

# ---------- Cross-platform open helper ----------
import platform, subprocess

def open_file_crossplatform(path):
    if not os.path.exists(path): return False
    system = platform.system()
    try:
        if system == "Windows": os.startfile(path)
        elif system == "Darwin": subprocess.run(["open", path])
        else: subprocess.run(["xdg-open", path])
        return True
    except:
        return False

import streamlit as st

def get_sheet_cached(sheet_name):
    """
    Read a sheet from Excel but cache in session_state to avoid losing changes
    """
    if "excel_cache" not in st.session_state:
        st.session_state["excel_cache"] = {}
    
    cache = st.session_state["excel_cache"]

    if sheet_name in cache:
        return cache[sheet_name]

    # Otherwise read from Excel
    df = read_sheet_safe(sheet_name)
    cache[sheet_name] = df
    return df

def update_sheet_cached(sheet_name, df):
    """
    Update session_state cache AND write to Excel safely
    """
    if "excel_cache" not in st.session_state:
        st.session_state["excel_cache"] = {}

    st.session_state["excel_cache"][sheet_name] = df

    # Save all sheets to Excel
    sheets_to_save = {
        SHEET_MB: get_sheet_cached(SHEET_MB),
        SHEET_GW: get_sheet_cached(SHEET_GW),
        SHEET_SPECS: get_sheet_cached(SHEET_SPECS)
    }
    saved, alt = atomic_write_all(EXCEL, sheets_to_save)
    if saved:
        apply_excel_coloring_and_separator([SHEET_MB, SHEET_GW])
    return saved, alt

