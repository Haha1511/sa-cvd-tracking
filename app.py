import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import os
import tempfile
import seaborn as sns
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
from datetime import datetime, timedelta

# ------------------ Page config must be first ------------------
st.set_page_config(page_title="SA Machine Data - Test6", layout="wide")

# ✅ Import backend module and confirm correct file path
import sa, importlib
importlib.reload(sa)  # force reload every run so updated code is used


# ✅ Import required functions and constants
from sa import (
    ensure_workbook,
    get_available_holes_for_part,
    get_features_for_part,
    safe_hole_sort_key,
    add_measurement_rows,
    show_trend_df,
    delete_rows_by_indexes,
    open_excel_file,
    get_reference_image_path,
    show_reference_photos,
    EXCEL,
    DATA_COLS,
    get_specs_df,
    export_specs_for_vendor,
    open_file_crossplatform
)



# Load Excel only once
if "df_plot_date" not in st.session_state:
    try:
        df = pd.read_excel("test6.xlsx")
        # Ensure necessary columns exist
        for col in ["Hole", "Value", "Timestamp", "Nominal", "LSL", "USL"]:
            if col not in df.columns:
                df[col] = pd.NA
        # Convert Timestamp to datetime
        df["Timestamp"] = pd.to_datetime(df["Timestamp"], errors="coerce")
        st.session_state.df_plot_date = df
    except Exception as e:
        st.error(f"Failed to load Excel: {e}")
        st.session_state.df_plot_date = pd.DataFrame(
            columns=["Hole","Value","Timestamp","Nominal","LSL","USL"]
        )

df_plot_date = st.session_state.df_plot_date

# Ensure workbook exists (backend will create if missing)
ensure_workbook()


# Header (logo + title)
col_logo1, col_title, col_logo2 = st.columns([1, 6, 1])

with col_logo1:
    logo_path = "fuji.png.png"  
    if os.path.exists(logo_path):
        try:
            # Add margin-top to push the logo down
            st.markdown('<div style="margin-top: 20px;">', unsafe_allow_html=True)
            st.image(logo_path, width=180)  # Adjust width as needed
            st.markdown('</div>', unsafe_allow_html=True)
        except Exception:
            st.write("")
    else:
        # Fallback remote emblem
        st.markdown('<div style="margin-top: 20px;">', unsafe_allow_html=True)
        st.image(
            "https://upload.wikimedia.org/wikipedia/commons/thumb/c/c2/Fuji_Electric_emblem.svg/512px-Fuji_Electric_emblem.svg.png",
            width=200,
        )
        st.markdown('</div>', unsafe_allow_html=True)

with col_title:
    st.markdown("<h1 style='text-align:center;'>SA Machine Data — Web Interface (Test6)</h1>", unsafe_allow_html=True)

with col_logo2:
    st.write("")  # Empty column to balance layout

st.markdown("---")
st.markdown("""
<style>

/* ✅ Global Light Theme Safety */
html, body, .stApp {
    background-color: #f8f9fa !important;
    color: #000000 !important;
}

/* ------------------ BUTTON COLORS ------------------ */

/* ✅ SAVE / SUBMIT → Green */
.stButton > button:contains("Save"),
.stButton > button:contains("Submit"),
.stButton > button:contains("Add"),
.stButton > button:contains("Confirm") {
    background-color: #28a745 !important; /* Green */
    color: #ffffff !important;
    font-weight: 700 !important;
    border-radius: 8px !important;
    border: none !important;
}
.stButton > button:contains("Save"):hover,
.stButton > button:contains("Submit"):hover,
.stButton > button:contains("Add"):hover,
.stButton > button:contains("Confirm"):hover {
    background-color: #1e7e34 !important;
}

/* ✅ VIEW / EXPORT / DOWNLOAD / OPEN → Blue */
.stButton > button:contains("View"),
.stButton > button:contains("Export"),
.stButton > button:contains("Download"),
.stButton > button:contains("Open") {
    background-color: #007bff !important; /* Blue */
    color: #ffffff !important;
    font-weight: 700 !important;
    border-radius: 8px !important;
    border: none !important;
}
.stButton > button:contains("View"):hover,
.stButton > button:contains("Export"):hover,
.stButton > button:contains("Download"):hover,
.stButton > button:contains("Open"):hover {
    background-color: #0056b3 !important;
}

/* ✅ DELETE → Red */
.stButton > button:contains("Delete"),
.stButton > button:contains("Remove"),
.stButton > button:contains("Clear") {
    background-color: #dc3545 !important; /* Red */
    color: #ffffff !important;
    font-weight: 700 !important;
    border-radius: 8px !important;
    border: none !important;
}
.stButton > button:contains("Delete"):hover,
.stButton > button:contains("Remove"):hover,
.stButton > button:contains("Clear"):hover {
    background-color: #b02a37 !important;
}

/* ✅ No weird Streamlit blue border */
.stButton > button {
    box-shadow: none !important;
}

</style>
""", unsafe_allow_html=True)

# ---------------- GLOBAL INPUT HIGHLIGHT ENABLE ----------------
st.markdown("""
<style>
/* Neon highlight effect for any focused input */
input:focus, textarea:focus, select:focus {
    border: 2px solid #00eaff !important;
    box-shadow: 0 0 8px #00eaff !important;
    outline: none !important;
}

/* Highlight container box */
.active-box {
    background-color: #1f2937 !important;
    border-radius: 12px;
    padding: 10px;
    transition: 0.15s ease-in-out;
    border: 2px solid #00eaff !important;
}

/* Reset box style */
.highlight-wrapper {
    padding: 10px;
    border-radius: 12px;
}
</style>

<script>
document.addEventListener("DOMContentLoaded", function() {
    const iframes = window.parent.document.querySelectorAll("iframe");

    function activateHighlight() {
        iframes.forEach((iframe) => {
            const doc = iframe.contentDocument || iframe.contentWindow.document;

            // select ALL Streamlit input widgets
            const inputs = doc.querySelectorAll('input, textarea, select');

            inputs.forEach((inp) => {
                inp.addEventListener("focus", () => {
                    // remove highlight from all wrappers
                    doc.querySelectorAll('.highlight-wrapper').forEach(w => {
                        w.classList.remove('active-box');
                    });

                    // find parent container to highlight
                    let parent = inp.closest('div[data-testid="stTextInput"]')
                              || inp.closest('div[data-testid="stSelectbox"]')
                              || inp.closest('div[data-testid="column"]')
                              || inp.parentElement;

                    if (parent) {
                        parent.classList.add('active-box');
                    }
                });
            });
        });
    }

    setTimeout(activateHighlight, 900);
});
</script>
""", unsafe_allow_html=True)


# Initialize session state keys used by UI
if "last_saved" not in st.session_state:
    st.session_state["last_saved"] = None
if "analysis_cache" not in st.session_state:
    st.session_state["analysis_cache"] = {}  # store last analysis results per (part,hole,feat)

# Tabs (added new Specs tab)
tabs = st.tabs(
    ["📥 Add Measurement", "📈 Trend Chart", "📊 View & Manage Data", "📘 View Spec", "🖼️ Reference Hole Photo", "📁 Excel File","📝Overall Summary"]
)

# ---------------- AUTO CLEAR + BUTTON HIGHLIGHT FLAGS ----------------
if "reset_inputs" not in st.session_state:
    st.session_state.reset_inputs = False

if "save_flash" not in st.session_state:
    st.session_state.save_flash = False
    

# ------------------ TAB 0: Add Measurement ------------------
with tabs[0]:
    st.subheader("Add Measurement")

    st.info("💡 Please make sure to close the Excel file before performing any actions.")

    # ---------------- SESSION STATE FOR AUTO-CLEAR ----------------
    if "clear_meas" not in st.session_state:
        st.session_state.clear_meas = False
    if "form_key" not in st.session_state:
        st.session_state.form_key = "form_add_0"
    if "form_counter" not in st.session_state:
        st.session_state.form_counter = 1
    if "clear_active" not in st.session_state:
        st.session_state.clear_active = False
    if "pending_photos_list" not in st.session_state:
        st.session_state.pending_photos_list = []  # List to hold multiple photos
    if "photo_counter" not in st.session_state:
        st.session_state.photo_counter = 0  # Counter for unique keys

    current_form_key = st.session_state.form_key

    with st.form(current_form_key):

        st.markdown('<div class="highlight-wrapper">', unsafe_allow_html=True)
        part = st.selectbox("Part Type", ["Mixing Block", "Gas/Water Block"])
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="highlight-wrapper">', unsafe_allow_html=True)
        machine_choice = st.selectbox("Machine", ["SA01", "SA02", "SA03", "Other"])
        st.markdown('</div>', unsafe_allow_html=True)

        machine = machine_choice
        if machine_choice == "Other":
            st.markdown('<div class="highlight-wrapper">', unsafe_allow_html=True)
            machine = st.text_input("Enter Machine Name", value="Unknown")
            st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="highlight-wrapper">', unsafe_allow_html=True)
        chamber = st.selectbox("Chamber", ["", "A", "B", "C", "D"])
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="highlight-wrapper">', unsafe_allow_html=True)
        piece_id = st.text_input("Piece ID / Serial Number")
        st.markdown('</div>', unsafe_allow_html=True)

        # ---------------- TIMESTAMP ----------------
        st.markdown('<div class="highlight-wrapper">', unsafe_allow_html=True)
        measured_date = st.date_input("Measured Date", value=datetime.now())
        st.markdown('</div>', unsafe_allow_html=True)

        # ---------------- PART IN / OUT ----------------
        st.markdown('<div class="highlight-wrapper">', unsafe_allow_html=True)
        part_flow = st.selectbox("Part Status (IN = returned, OUT = sent)", ["IN", "OUT"])
        st.markdown('</div>', unsafe_allow_html=True)

        # ---------------- BATCH CLEANING NUMBER ----------------
        st.markdown('<div class="highlight-wrapper">', unsafe_allow_html=True)
        batch_number = st.text_input("Batch Cleaning Number (optional)")
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="highlight-wrapper">', unsafe_allow_html=True)
        notes = st.text_input("Notes (optional)")
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown("**Measurements (mm)** — leave empty to skip a field")

        holes = ["1", "2", "3", "4", "5"]
        cols_in = st.columns(3)
        inputs = []
        col_idx = 0
        for h in holes:
            with cols_in[col_idx]:

                st.markdown('<div class="highlight-wrapper">', unsafe_allow_html=True)
                st.markdown(f"**H{h}**")

                for f in get_features_for_part(part, h):
                    key = f"meas_H{h}_{f}"
                    default_value = "" if st.session_state.clear_meas else st.session_state.get(key, "")
                    val = st.text_input(f"{f}", key=key, value=default_value)
                    inputs.append({"Hole": str(h), "Feature": f, "Value": val})

                st.markdown('</div>', unsafe_allow_html=True)

            col_idx = (col_idx + 1) % 3

        # ======================================================
        # 📸 MULTIPLE PHOTO SECTION (Unlimited Uploads)
        # ======================================================
        st.markdown("---")
        st.subheader("📸 Upload Photos for Specific Holes (Multiple Allowed)")

        idx = st.session_state.photo_counter

        photo_hole = st.selectbox(
            "Select Hole",
            ["1", "2", "3", "4", "5"],
            key=f"photo_hole_add_{idx}"
        )
        photo_feature = st.selectbox(
            "Select Feature",
            ["Inner", "Outer"],
            key=f"photo_feature_add_{idx}"
        )
        uploaded_photo = st.file_uploader(
            "Upload Photo (PNG/JPG)",
            type=["png", "jpg", "jpeg"],
            key=f"photo_upload_add_{idx}"
        )

        # Add photo to pending list
        if uploaded_photo:
            st.session_state.pending_photos_list.append({
                "hole": photo_hole,
                "feature": photo_feature,
                "file": uploaded_photo
            })
            st.success(f"Added photo for H{photo_hole} - {photo_feature}")

        # Button to add another photo (increments counter)
        if st.form_submit_button("➕ Add Another Photo"):
            st.session_state.photo_counter += 1
            st.rerun()

        # Show all pending photos
        if st.session_state.pending_photos_list:
            st.markdown("**Pending Photos:**")
            for idx, p in enumerate(st.session_state.pending_photos_list, start=1):
                st.write(f"{idx}. Hole: H{p['hole']} - {p['feature']}")

        # ================= Form submit button =================
        submitted = st.form_submit_button("Save Measurements")

    # ---------------- AFTER FORM CREATION ----------------
    if st.session_state.get("clear_active", False):
        st.session_state.clear_meas = False
        st.session_state.clear_active = False

    # ---------------- AFTER SUBMIT LOGIC ----------------
    if submitted:
        measurements = []
        for it in inputs:
            raw = str(it["Value"]).strip()
            if raw == "" or raw == "-":
                continue
            try:
                measurements.append({"Hole": it["Hole"], "Feature": it["Feature"], "Value": float(raw)})
            except Exception:
                st.warning(f"Invalid number for H{it['Hole']} {it['Feature']}: '{raw}' — skipped")

        # Attach all pending photos to corresponding measurements
        for p in st.session_state.pending_photos_list:
            hole = p["hole"]
            feature = p["feature"]
            file = p["file"]

            os.makedirs("uploaded_images", exist_ok=True)
            img_filename = f"{piece_id}_H{hole}_{feature}.jpg"
            img_path = os.path.join("uploaded_images", img_filename)

            with open(img_path, "wb") as f:
                f.write(file.getbuffer())

            # Assign the image path to all matching measurements
            for m in measurements:
                if m["Hole"] == hole and m["Feature"].lower() == feature.lower():
                    m["ImagePath"] = img_path

        st.session_state.pending_photos_list = []  # clear after attaching
        st.session_state.photo_counter = 0  # reset counter

        if not piece_id:
            st.error("Piece ID / Serial Number is required.")
        elif not measurements:
            st.info("No valid measurements entered; nothing saved.")
        else:
            # ✅ Pass measured_date and batch_number to add_measurement_rows
            ok, msg = add_measurement_rows(
                part, machine, chamber, piece_id, part_flow, notes, 
                measurements, measured_date=measured_date, batch_number=batch_number
            )

            if ok:
                st.session_state.clear_meas = True
                st.session_state.clear_active = True

                now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                st.session_state["last_saved"] = now

                st.session_state.form_key = f"form_add_{st.session_state.form_counter}"
                st.session_state.form_counter += 1

                # IMPORTANT: only store message here (do NOT display it yet)
                st.session_state["pending_success"] = f"✅ Saved to Excel. ({msg}) — {now}"

                if "analysis_cache" in st.session_state:
                    st.session_state["analysis_cache"].pop((part,), None)

                st.rerun()
            else:
                st.error(f"❌ Failed to save ***PLEASE MAKE SURE CLOSED EXCEL FILE FIRST***: {msg}")

    # ---------------- SHOW SUCCESS MESSAGE AT BOTTOM ----------------
    if "pending_success" in st.session_state:
        st.success(st.session_state["pending_success"])
        del st.session_state["pending_success"]


# ------------------ TAB 1: Trend Chart (with Analysis) ------------------
with tabs[1]:
    st.subheader("Trend Chart")
    st.info("💡 Please make sure to close the Excel file before performing any actions.")

    auto = st.checkbox("Auto Refresh (30s)")
    if auto:
        st.markdown('<meta http-equiv="refresh" content="30">', unsafe_allow_html=True)

    part_trend = st.selectbox("Part Type", ["Mixing Block", "Gas/Water Block"], key="trend_part")
    df_trend = show_trend_df(part_trend)

    if df_trend is None or df_trend.empty:
        st.info("No measurement data available yet. Add measurements in Add Measurement tab.")
    else:
        # Use Measured Date instead of Timestamp
        df_trend["Measured Date"] = pd.to_datetime(df_trend["Measured Date"], errors="coerce").dt.date
        df_trend = df_trend.dropna(subset=["Measured Date"])

        # Machine filter
        machines = ["All"] + sorted(df_trend["Machine"].dropna().unique().tolist())
        mch = st.selectbox("Machine", machines)
        if mch != "All":
            df_trend = df_trend[df_trend["Machine"] == mch]

        # Chamber filter
        chambers = ["All"] + sorted(df_trend["Chamber"].dropna().unique().tolist())
        ch = st.selectbox("Chamber", chambers)
        if ch != "All":
            df_trend = df_trend[df_trend["Chamber"] == ch]

        # ---- PIECE ID FILTER ADDITION ----
        piece_ids = ["All"] + sorted(df_trend["Piece ID"].dropna().unique().tolist())
        pid = st.selectbox("Piece ID", piece_ids)
        if pid != "All":
            df_trend = df_trend[df_trend["Piece ID"] == pid]

        # Hole selection (multi-select)
        raw_holes = df_trend["Hole"].dropna().unique().tolist()
        valid_holes = [h for h in raw_holes if str(h).strip().upper().startswith("H") and str(h).strip()[1:].isdigit()]
        holes = sorted(valid_holes, key=safe_hole_sort_key)
        if part_trend.lower() == "gas/water block":
            holes = [f"H{i}" for i in range(1, 6)]
        selected_holes = st.multiselect("Select Hole(s)", holes, default=[holes[0]])

        # Feature selection
        features = sorted(df_trend[df_trend["Hole"].isin(selected_holes)]["Feature"].dropna().unique().tolist())
        feat = st.selectbox("Feature", features)

        # Filter df for selected holes & feature
        df_plot_filtered = df_trend[(df_trend["Hole"].isin(selected_holes)) & (df_trend["Feature"] == feat)].sort_values("Measured Date")

        if df_plot_filtered.empty:
            st.warning("No data for this selection.")
        else:
            # ------------------ Trend Data Table ------------------
            st.markdown("### 📘 Trend Data Table")
            if len(selected_holes) == 1:
                # Single hole → show spec coloring
                single_hole = selected_holes[0]
                df_single = df_plot_filtered[df_plot_filtered["Hole"] == single_hole]

                def highlight_value_only(row):
                    try:
                        val = float(row["Value"])
                    except:
                        val = None
                    try:
                        lsl = float(row["LSL"]) if "LSL" in row and not pd.isna(row["LSL"]) else None
                    except:
                        lsl = None
                    try:
                        usl = float(row["USL"]) if "USL" in row and not pd.isna(row["USL"]) else None
                    except:
                        usl = None
                    colors = [""] * len(row)
                    value_idx = row.index.get_loc("Value")
                    if val is not None:
                        if (lsl is not None and val < lsl) or (usl is not None and val > usl):
                            colors[value_idx] = "background-color: #FFCCCC"  # RED → out of spec
                        else:
                            colors[value_idx] = "background-color: #CCFFCC"  # GREEN → within spec
                    return colors

                st.dataframe(
                    df_single[["Timestamp", "Measured Date", "Machine", "Chamber", "Piece ID", "Hole", "Feature", "Value", "LSL", "USL"]].style.apply(highlight_value_only, axis=1),
                    use_container_width=True
                )

            # ------------------ Date Range Filter ------------------
            min_dt = df_plot_filtered["Measured Date"].min()
            max_dt = df_plot_filtered["Measured Date"].max()

            if pd.isna(min_dt) or pd.isna(max_dt):
                st.warning("⚠ No valid date values found.")
                df_plot_date = df_plot_filtered
            else:
                if min_dt == max_dt:
                    st.info(f"Only one date found: **{min_dt}**. No date filter applied.")
                    df_plot_date = df_plot_filtered
                else:
                    col_d1, col_d2 = st.columns(2)
                    with col_d1:
                        start_date = st.date_input("Start Measured Date", value=min_dt, min_value=min_dt, max_value=max_dt)
                    with col_d2:
                        end_date = st.date_input("End Measured Date", value=max_dt, min_value=min_dt, max_value=max_dt)

                    if start_date > end_date:
                        st.error("❌ Start measured date cannot be after end measured date.")
                        df_plot_date = df_plot_filtered
                    else:
                        df_plot_date = df_plot_filtered[(df_plot_filtered["Measured Date"] >= start_date) &
                                                        (df_plot_filtered["Measured Date"] <= end_date)]

                if df_plot_date.empty:
                    st.warning("⚠ No data in the selected date range.")
                    
                # ------------------ Professional Plotting ------------------
                fig, ax = plt.subplots(figsize=(10, 4.5))
                ax.set_facecolor("#f5f5f5")
                ax.grid(True, linestyle="--", linewidth=0.7, alpha=0.5)

                hole_colors_dict = {}
                default_colors = ["#1F77B4", "#FF5733", "#33FF57", "#9B59B6",
                                  "#F1C40F", "#E67E22", "#1ABC9C", "#8E44AD"]
                for i, hole in enumerate(selected_holes):
                    hole_colors_dict[hole] = default_colors[i % len(default_colors)]

                # REMOVE spec_shaded — draw spec zone for each hole individually

                # Keep nominal/LSL/USL flags for drawing only once per chart
                nominal_drawn = False
                lsl_drawn = False
                usl_drawn = False

                for i, hole in enumerate(selected_holes):
                    df_h = df_plot_date[df_plot_date["Hole"] == hole]
                    if df_h.empty:
                        continue

                    x_values = range(1, len(df_h)+1)
                    y_values = df_h["Value"].astype(float)

                    # ✅ Draw spec zone for this hole
                    lsl_val = df_h["LSL"].dropna().iloc[0] if "LSL" in df_h.columns and not df_h["LSL"].dropna().empty else None
                    usl_val = df_h["USL"].dropna().iloc[0] if "USL" in df_h.columns and not df_h["USL"].dropna().empty else None
                    if lsl_val is not None and usl_val is not None:
                        ax.fill_between(x_values, lsl_val, usl_val, color="#d4f4dd", alpha=0.3, label=f"{hole} Spec zone")

                    # Plot values & markers
                    ax.plot(x_values, y_values, color=hole_colors_dict[hole], linewidth=2.2, label=f"{hole} Value")
                    ax.scatter(x_values, y_values, color=hole_colors_dict[hole], s=65,
                               edgecolors="white", linewidth=0.7, zorder=4)

                    # Trend line
                    if len(x_values) >= 2:
                        if len(set(y_values)) == 1:
                            ax.plot(
                                x_values,
                                [y_values.iloc[0]] * len(x_values),
                                linestyle="--",
                                color=hole_colors_dict[hole],
                                linewidth=2,
                                alpha=0.7,
                                label=f"{hole} Trend"
                            )
                        else:
                            z = np.polyfit(list(x_values), y_values, 1)
                            p = np.poly1d(z)
                            ax.plot(
                                x_values,
                                p(x_values),
                                linestyle="--",
                                color=hole_colors_dict[hole],
                                linewidth=2,
                                alpha=0.8,
                                zorder=3,
                                label=f"{hole} Trend"
                            )

                    # Annotate last value
                    ax.text(
                        x_values[-1] + 0.1,
                        y_values.iloc[-1],
                        f"{y_values.iloc[-1]:.2f}",
                        fontsize=9,
                        fontweight="bold",
                        color=hole_colors_dict[hole],
                        va="bottom",
                        ha="left"
                    )

                    # Nominal/LSL/USL lines — draw only once
                    nominal_val = df_h["Nominal"].dropna().iloc[0] if "Nominal" in df_h.columns and not df_h["Nominal"].dropna().empty else None
                    if nominal_val is not None and not nominal_drawn:
                        ax.axhline(nominal_val, linestyle="--", color="yellow", linewidth=1.5, alpha=0.7, label="Nominal")
                        nominal_drawn = True

                    if lsl_val is not None and not lsl_drawn:
                        ax.axhline(lsl_val, linestyle="--", color="red", linewidth=1.5, alpha=0.7, label="LSL")
                        lsl_drawn = True

                    if usl_val is not None and not usl_drawn:
                        ax.axhline(usl_val, linestyle="--", color="green", linewidth=1.5, alpha=0.7, label="USL")
                        usl_drawn = True

                ax.set_xlabel("Measurement Count", fontsize=10, color="#333333")
                ax.set_ylabel("Measurement (mm)", fontsize=10, color="#333333")
                ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, _: f"{x:.2f} mm"))
                ax.tick_params(colors="#333333", labelsize=9)
                mach_label = mch if mch != "All" else "All Machines"
                ch_label = ch if ch != "All" else "All Chambers"
                ax.set_title(
                    f"Trend — {part_trend} ({feat})\n{mach_label}, {ch_label}",
                    fontsize=13,
                    color="#222222",
                    fontweight="bold"
                )

                ax.legend(fontsize=9, loc="upper left", framealpha=0.9)

                plt.tight_layout()
                st.pyplot(fig)

                # --- Download ---
                buf = BytesIO()
                fig.savefig(buf, format="png", dpi=300, bbox_inches="tight")

                st.download_button(
                    "📥 Download Trend Chart",
                    buf.getvalue(),
                    file_name=f"Trend_{feat}.png",
                    mime="image/png",
                )

                # ------------------ Trend Analysis UI (Multi-hole with specs) ------------------
                st.markdown(
                    """
                    <style>
                    .status-green { color: #28a745; font-weight:700; }
                    .status-green2 { color: #2ecc71; font-weight:700; }
                    .status-yellow { color: #f1c40f; font-weight:700; }
                    .status-red { color: #e74c3c; font-weight:700; }
                    .analysis-box { background:#f2f2f2; padding:12px; border-radius:8px; margin-bottom:10px; color:#000000; }
                    .analysis-title { font-weight:800; font-size:1.02rem; }
                    .last-val-badge { font-weight:800; padding:4px 8px; border-radius:6px; color:#fff; }
                    .last-val-green { background-color:#28a745; }
                    .last-val-yellow { background-color:#f1c40f; color:#000; }
                    .last-val-red { background-color:#e74c3c; }
                    </style>
                    """,
                    unsafe_allow_html=True,
                )

                col_an1, col_an2 = st.columns([3, 1])
                with col_an1:
                    st.write("Click **Analyze Trend** to compute linear trend (slope), R², Δ, last value, and spec info per hole.")
                with col_an2:
                    analyze_btn = st.button("Analyze Trend", key=f"analyze_{part_trend}_{feat}")

                # Initialize cache
                if "analysis_cache" not in st.session_state:
                    st.session_state["analysis_cache"] = {}

                def compute_trend_analysis(df_in):
                    result = {
                        "slope": 0.0, "r2": 0.0, "delta": 0.0,
                        "trend_status": ("Stable", "status-green"),
                        "prox_status": None, "nominal": None,
                        "lsl": None, "usl": None, "last_val": None,
                    }

                    if df_in is None or df_in.empty:
                        return result

                    df = df_in.sort_values("Timestamp").copy()
                    y = pd.to_numeric(df["Value"], errors="coerce").fillna(0).astype(float).values
                    if len(y) == 0: return result

                    x = (df["Timestamp"] - df["Timestamp"].min()).dt.total_seconds() / 86400.0
                    if len(x) >= 2 and not np.allclose(y, y[0]):
                        p = np.polyfit(x, y, 1)
                        slope, intercept = float(p[0]), float(p[1])
                        yhat = np.polyval(p, x)
                        denom = np.sum((y - np.mean(y)) ** 2)
                        r2 = 1 - np.sum((y - yhat) ** 2) / denom if denom != 0 else 1.0
                    else:
                        slope = 0.0; intercept = float(y[0]); r2 = 0.0

                    delta = float(y[-1] - y[0])
                    # --- FIXED TREND LOGIC (no abs(), directional and correct) ---
                    if slope > 0.1:
                        trend_status = ("Rapid upward change", "status-red")
                    elif slope < -0.1:
                        trend_status = ("Rapid downward change", "status-red")
                    elif slope > 0.01:
                        trend_status = ("Drifting upward", "status-yellow")
                    elif slope < -0.01:
                        trend_status = ("Drifting downward", "status-yellow")
                    else:
                        trend_status = ("Stable", "status-green")

                    # Nominal/LSL/USL
                    nominal = df["Nominal"].dropna().iloc[0] if "Nominal" in df.columns and not df["Nominal"].dropna().empty else None
                    lsl = df["LSL"].dropna().iloc[0] if "LSL" in df.columns and not df["LSL"].dropna().empty else None
                    usl = df["USL"].dropna().iloc[0] if "USL" in df.columns and not df["USL"].dropna().empty else None
                    last_val = float(y[-1])

                    prox_status = None
                    if lsl is not None and usl is not None:
                        if last_val < lsl or last_val > usl:
                            prox_status = ("Out of spec", "status-red", f"Last value {last_val:.4f} outside spec [{lsl}, {usl}]")
                        else:
                            span = usl - lsl if (usl - lsl) != 0 else 1.0
                            dist_to_nearest = min(abs(last_val - lsl), abs(usl - last_val))
                            proximity = dist_to_nearest / span
                            if proximity < 0.10:
                                prox_status = ("Near limit", "status-yellow", f"Last value {last_val:.4f} within 10% of limit")
                            else:
                                prox_status = ("Within spec", "status-green2", f"Last value {last_val:.4f} comfortably within spec")

                    return {
                        "slope": slope, "r2": r2, "delta": delta,
                        "trend_status": trend_status, "prox_status": prox_status,
                        "nominal": nominal, "lsl": lsl, "usl": usl,
                        "last_val": last_val,
                    }

                # Loop through selected holes
                for hole in selected_holes:
                    df_hole = df_plot_date[df_plot_date["Hole"] == hole]
                    cache_key = (part_trend, hole, feat)
                    result = None

                    if analyze_btn:
                        try:
                            result = compute_trend_analysis(df_hole)
                            st.session_state["analysis_cache"][cache_key] = result
                        except Exception as e:
                            st.error(f"Analysis failed for {hole}: {e}")
                            result = None
                    elif cache_key in st.session_state["analysis_cache"]:
                        result = st.session_state["analysis_cache"][cache_key]

                    if result:
                        trend_label, trend_css = result["trend_status"]
                        prox = result.get("prox_status")
                        summary_css = prox[1] if prox else trend_css
                        last_val = result.get("last_val")
                        last_badge_class = "last-val-green"
                        if prox:
                            if prox[1] == "status-red": last_badge_class = "last-val-red"
                            elif prox[1] == "status-yellow": last_badge_class = "last-val-yellow"

                        summary_html = (
                            f"<div class='analysis-box'>"
                            f"<div class='analysis-title'>"
                            f"<span class='{summary_css}'>"
                            f"<b>{hole}</b> — trend: <b>{trend_label.lower()}</b>, slope = <code>{result['slope']:.4f}</code> mm/day, R² = <code>{result['r2']:.3f}</code>, Δ = <code>{result['delta']:.4f}</code>."
                            f"</span></div>"
                            f"<div style='margin-top:4px;'>"
                            f"<b>Nominal:</b> {result.get('nominal','N/A')} &nbsp;&nbsp; "
                            f"<b>LSL:</b> {result.get('lsl','N/A')} &nbsp;&nbsp; "
                            f"<b>USL:</b> {result.get('usl','N/A')}"
                            f"</div>"
                            f"<span class='last-val-badge {last_badge_class}'>Last: {last_val:.4f}</span>"
                        )

                        if prox:
                            summary_html += (
                                f"<div style='margin-top:4px;'>"
                                f"<b>Spec Check:</b> <span class='{prox[1]}'>{prox[0]}</span><br>"
                                f"<i>{prox[2]}</i></div>"
                            )

                        summary_html += "</div>"
                        st.markdown(summary_html, unsafe_allow_html=True)



    with st.expander("ℹ️ Explain Trend Metrics"):
        st.markdown(
            """
            ### 📘 What Each Metric Means

            **🔹 Slope (mm/day)**  
            - Shows how fast the measurement is increasing or decreasing  
            - Positive slope → rising  
            - Negative slope → falling  
            - Near zero → stable

            **🔹 R² (R-squared)**  
            - Measures how well the data fits a straight-line trend  
            - Range: **0 to 1**  
            - **1.0 = perfect trend line**  
            - **0.0 = no trend pattern (random / stable)**  
            - Helps judge stability, drift, or sudden changes

            **🔹 Δ (first → last)**  
            - Total change between first and last measurements  
            - Shows overall shift regardless of slope

            **🔹 Last Value**  
            - Latest measurement  
            - Automatically color-coded (green/yellow/red)

            **🔹 Nominal / LSL / USL**  
            - Nominal = ideal/target value  
            - LSL = Lower limit  
            - USL = Upper limit  
            - Used for spec-check

            **🔹 Spec Check**  
            - 🟢 Within spec  
            - 🟡 Near limit  
            - 🔴 Out of spec  
            """
        )
    # ------------------ Export & Update Trend Charts in Excel (Full) ------------------
    import os
    from io import BytesIO
    import matplotlib.pyplot as plt
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import PatternFill
    from matplotlib.patches import Rectangle
    import streamlit as st
    import numpy as np
    import openpyxl
    from openpyxl.utils import get_column_letter

    TREND_EXCEL = "trendchart.xlsx"

    # ------------------ AUTO ADJUST COLUMNS FUNCTION ------------------
    def auto_adjust_columns_excel(wb, sheet_names):
        """
        Adjust column widths for given sheets in an openpyxl workbook
        """
        for sheet_name in sheet_names:
            if sheet_name not in wb.sheetnames:
                continue
            sheet = wb[sheet_name]

            for i, col_cells in enumerate(sheet.iter_cols(1, sheet.max_column), start=1):
                col_letter = get_column_letter(i)
                max_length = 0
                for cell in col_cells:
                    if cell.value is None:
                        continue
                    cell_len = len(str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len

                # Special adjustments for known long columns
                header = sheet.cell(row=1, column=i).value
                if header in ["Notes", "Image Path", "Batch Cleaning"]:
                    adjusted_width = max(25, max_length + 2)  # long text
                elif header in ["Measured Date", "Part In/Out", "Machine", "Part Type", "Chamber", "Hole", "Feature", "Status"]:
                    adjusted_width = max(12, max_length + 2)
                else:
                    adjusted_width = max(15, max_length + 2)

                sheet.column_dimensions[col_letter].width = min(adjusted_width, 50)  # cap max width

    # ---------------- Update Trend Charts Button ----------------
    if st.button("Update Trend Charts"):
        if df_trend is None or df_trend.empty:
            st.warning("No measurement data available to generate trend charts.")
        else:
            try:
                # Initialize progress bar and status text
                progress = st.progress(0)
                status_text = st.empty()

                # Create new workbook
                wb = Workbook()
                wb.remove(wb.active)

                parts = ["Mixing Block", "Gas/Water Block"]

                # Calculate total steps for progress
                total_steps = 0
                for part in parts:
                    df_part_tmp = show_trend_df(part)
                    if df_part_tmp.empty:
                        continue
                    machines_tmp = sorted(df_part_tmp["Machine"].dropna().unique().tolist())
                    chambers_tmp = sorted(df_part_tmp["Chamber"].dropna().unique().tolist())
                    piece_ids_tmp = sorted(df_part_tmp["Piece ID"].dropna().unique().tolist())
                    total_steps += len(machines_tmp) * len(chambers_tmp) * len(piece_ids_tmp)
                if total_steps == 0:
                    total_steps = 1  # avoid division by zero
                current_step = 0

                for part in parts:
                    df_part = show_trend_df(part)
                    if df_part.empty:
                        continue

                    machines = sorted(df_part["Machine"].dropna().unique().tolist())
                    chambers = sorted(df_part["Chamber"].dropna().unique().tolist())
                    piece_ids_list = sorted(df_part["Piece ID"].dropna().unique().tolist())
                    combos = [("All", "All", "All")] + [(m, "All", "All") for m in machines] + [(m, ch, "All") for m in machines for ch in chambers] + [(m, ch, pid) for m in machines for ch in chambers for pid in piece_ids_list]

                    for mach, ch, pid in combos:
                        df_filtered = df_part.copy()
                        if mach != "All":
                            df_filtered = df_filtered[df_filtered["Machine"] == mach]
                        if ch != "All":
                            df_filtered = df_filtered[df_filtered["Chamber"] == ch]
                        if pid != "All":
                            df_filtered = df_filtered[df_filtered["Piece ID"] == pid]

                        if df_filtered.empty:
                            current_step += 1
                            progress.progress(min(current_step / total_steps, 1.0))
                            continue

                        # Update status
                        status_text.text(f"Processing {part} | Machine: {mach} | Chamber: {ch} | Piece ID: {pid}...")
                        current_step += 1
                        progress.progress(min(current_step / total_steps, 1.0))

                        # Sheet name: include Part, Machine, Chamber, PieceID
                        sheet_name = f"{part}_{mach}_{ch}_{pid}".replace(" ", "_").replace("/", "_")[:31]
                        sheet = wb.create_sheet(title=sheet_name)
                        current_row = 1

                        holes = sorted(df_filtered["Hole"].dropna().unique().tolist(), key=safe_hole_sort_key)
                        default_colors = ["#1F77B4", "#FF5733", "#33FF57", "#9B59B6",
                                          "#F1C40F", "#E67E22", "#1ABC9C", "#8E44AD"]
                        hole_colors_dict = {h: default_colors[i % len(default_colors)] for i, h in enumerate(holes)}

                        for hole in holes:
                            feats = sorted(df_filtered[df_filtered["Hole"] == hole]["Feature"].dropna().unique().tolist())
                            for feat in feats:
                                dfp = df_filtered[(df_filtered["Hole"] == hole) & (df_filtered["Feature"] == feat)].sort_values("Timestamp")
                                if dfp.empty:
                                    continue

                                # -------- WRITE TABLE WITH PASS/FAIL COLORS --------
                                for c_idx, col in enumerate(dfp.columns, start=1):
                                    sheet.cell(row=current_row, column=c_idx, value=col)

                                for r_i, r in enumerate(dfp.itertuples(index=False, name=None), start=current_row+1):
                                    for c_i, val in enumerate(r, start=1):
                                        cell = sheet.cell(row=r_i, column=c_i, value=val)
                                        if dfp.columns[c_i-1] == "Value":
                                            try:
                                                val_f = float(val)
                                                lsl = float(dfp["LSL"].iloc[r_i - current_row - 1]) if "LSL" in dfp.columns else None
                                                usl = float(dfp["USL"].iloc[r_i - current_row - 1]) if "USL" in dfp.columns else None
                                                if (lsl is not None and val_f < lsl) or (usl is not None and val_f > usl):
                                                    cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                                                else:
                                                    cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                                            except:
                                                pass

                                table_rows = len(dfp) + 1
                                current_row += table_rows + 1

                                # -------- CHART (match website style exactly) --------
                                fig_height = max(4, table_rows * 0.35)
                                fig_width = 10
                                fig, ax = plt.subplots(figsize=(fig_width, fig_height))
                                ax.set_facecolor("#f5f5f5")
                                ax.grid(True, linestyle="--", linewidth=0.7, alpha=0.5)

                                x_values = list(range(1, len(dfp)+1))
                                y_values = dfp["Value"].astype(float)
                                color = hole_colors_dict[hole]

                                ax.plot(x_values, y_values, color=color, linewidth=2.2, label=f"{hole} Value", zorder=3)
                                ax.scatter(x_values, y_values, color=color, s=65, edgecolors="white", linewidth=0.7, zorder=4)

                                if len(x_values) >= 2:
                                    z = np.polyfit(x_values, y_values, 1)
                                    p = np.poly1d(z)
                                    ax.plot(x_values, p(x_values), linestyle="--", color=color, alpha=0.7, label=f"{hole} Trend")

                                nominal_val = dfp["Nominal"].dropna().iloc[0] if "Nominal" in dfp.columns and not dfp["Nominal"].dropna().empty else None
                                lsl_val = dfp["LSL"].dropna().iloc[0] if "LSL" in dfp.columns and not dfp["LSL"].dropna().empty else None
                                usl_val = dfp["USL"].dropna().iloc[0] if "USL" in dfp.columns and not dfp["USL"].dropna().empty else None

                                if nominal_val is not None:
                                    ax.axhline(nominal_val, linestyle="--", color="yellow", linewidth=1.5, alpha=0.7)
                                if lsl_val is not None:
                                    ax.axhline(lsl_val, linestyle="--", color="red", linewidth=1.5, alpha=0.7)
                                if usl_val is not None:
                                    ax.axhline(usl_val, linestyle="--", color="green", linewidth=1.5, alpha=0.7)

                                if lsl_val is not None and usl_val is not None and len(x_values) > 0:
                                    rect_x0 = x_values[0] - 0.5
                                    rect_width = x_values[-1] - x_values[0] + 1.0
                                    rect_y0 = lsl_val
                                    rect_height = usl_val - lsl_val
                                    rect = Rectangle((rect_x0, rect_y0), rect_width, rect_height, color="#d4f4dd", alpha=0.35, zorder=1, ec="none")
                                    ax.add_patch(rect)

                                ax.text(x_values[-1]+0.1, y_values.iloc[-1], f"{y_values.iloc[-1]:.2f}",
                                        fontsize=9, fontweight="bold", color=color, va="bottom", ha="left", zorder=5)
                                ax.set_xlabel("Measurement Count", fontsize=10, color="#333333")
                                ax.set_ylabel("Measurement (mm)", fontsize=10, color="#333333")
                                ax.tick_params(colors="#333333", labelsize=9)
                                # Include Piece ID in chart title
                                ax.set_title(f"{part} | {hole} | {feat} | Piece ID: {pid}", fontsize=12, color="#222222")
                                ax.legend(fontsize=8, loc="upper left")
                                plt.tight_layout()

                                # Save chart to buffer & insert into Excel
                                chart_buf = BytesIO()
                                fig.savefig(chart_buf, dpi=200, format="png")
                                plt.close(fig)
                                chart_buf.seek(0)
                                img = XLImage(chart_buf)
                                img.width = fig_width * 90
                                img.height = fig_height * 90
                                img.anchor = f"A{current_row}"
                                sheet.add_image(img)
                                current_row += 5 + int(fig_height*4)

                # ------------------ AUTO ADJUST ALL COLUMNS ------------------
                try:
                    auto_adjust_columns_excel(wb, wb.sheetnames)
                except Exception as e:
                    st.warning(f"Auto-adjust columns failed: {e}")

                # --- Save workbook to disk ---
                wb.save(TREND_EXCEL)

                # --- Prepare download buffer ---
                excel_buffer = BytesIO()
                wb_for_buffer = load_workbook(TREND_EXCEL)
                wb_for_buffer.save(excel_buffer)
                excel_buffer.seek(0)

                st.download_button(
                    label="📥 Download Trend Charts (Excel)",
                    data=excel_buffer,
                    file_name="trendchart.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                progress.progress(1.0)
                status_text.text("✅ Completed")
                st.success("✅ Trend charts updated successfully!")

            except Exception as e:
                st.error(f"Failed to update trend charts: {e}")

    # ---------------- Open Trend Charts Button ----------------
    if st.button("Open Trend Charts"):
        trend_path = os.path.abspath(TREND_EXCEL)
        if os.path.exists(trend_path):
            try:
                os.startfile(trend_path)  # Windows only
                st.info(f"Excel file opened directly: {trend_path}")
            except Exception as e:
                st.warning(f"Could not open Excel automatically: {e}")
                st.info(f"Please manually open '{trend_path}' from the project folder.")
        else:
            st.warning("The trendchart.xlsx file does not exist yet. Please update the charts first.")

    # Define the path to save images
    IMAGE_DIR = "uploaded_images"
    os.makedirs(IMAGE_DIR, exist_ok=True)

# ------------------ TAB 2: View & Manage Data ------------------
with tabs[2]:
    st.subheader("View & Manage Data")
    st.info("💡 Please make sure to close the Excel file before performing any actions.")

    # ------------------ CONFIG PATHS ------------------
    import os
    import pathlib

    # Make EXCEL path flexible
    EXCEL = st.session_state.get("EXCEL_PATH", "test6.xlsx")  # default Excel filename
    # Ensure uploaded_images folder exists
    IMG_DIR = pathlib.Path("uploaded_images")
    IMG_DIR.mkdir(exist_ok=True)

    # Toast handler
    if "toast_msg" in st.session_state:
        st.toast(st.session_state["toast_msg"], icon=st.session_state.get("toast_icon", "ℹ️"))
        del st.session_state["toast_msg"]
        if "toast_icon" in st.session_state:
            del st.session_state["toast_icon"]

    part_view = st.selectbox("Which part to view", ["Mixing Block", "Gas/Water Block"], key="view_part")
    sheet_name = "Mixing Block Data" if part_view == "Mixing Block" else "Gas-Water Block Data"

    # --- Read Excel ---
    try:
        df_view_orig = pd.read_excel(EXCEL, sheet_name=sheet_name)
        df_view_orig = df_view_orig.loc[:, ~df_view_orig.columns.str.contains("^Unnamed", case=False)]

        if "Piece ID" in df_view_orig.columns and "Part In/Out" in df_view_orig.columns:
            cols = list(df_view_orig.columns)
            if "Batch Cleaning" in cols:
                cols.remove("Batch Cleaning")
            cols.remove("Part In/Out")
            idx = cols.index("Piece ID")
            cols.insert(idx + 1, "Part In/Out")
            cols.insert(idx + 2, "Batch Cleaning")
            df_view_orig = df_view_orig[cols]

        if "Timestamp" in df_view_orig.columns and "Measured Date" in df_view_orig.columns:
            cols = list(df_view_orig.columns)
            cols.remove("Measured Date")
            idx = cols.index("Timestamp")
            cols.insert(idx + 1, "Measured Date")
            df_view_orig = df_view_orig[cols]

    except PermissionError:
        st.error("❌ Excel file is open. Please close it first.")
        st.stop()
    except Exception as e:
        st.error(f"⚠️ Failed to load data: {e}")
        df_view_orig = pd.DataFrame()

    # ------------------ ENSURE df_view IS ALWAYS DEFINED ------------------
    if df_view_orig.empty:
        df_view = pd.DataFrame()  # empty DataFrame if Excel is empty or missing
    else:
        df_view = df_view_orig.copy()
        df_view["_orig_index"] = df_view.index  # Track original Excel row

        # Safe conversion for Measured Date
        if "Measured Date" in df_view.columns:
            df_view["Measured Date"] = pd.to_datetime(df_view["Measured Date"], errors="coerce")

        # Reset Filters button
        if st.button("🔄 Reset Filters"):
            for key in ["Machine", "Chamber", "Hole", "Feature", "Status", "Piece_ID_filter"]:
                if key in st.session_state:
                    del st.session_state[key]
            st.rerun()

        # Collapsible filter panel
        with st.expander("Show / Hide Filters"):
            col1, col2, col3 = st.columns(3)
            col4, col5, col6 = st.columns(3)

            # ---------------- MACHINE FILTER ----------------
            machine_options = sorted(df_view_orig["Machine"].dropna().unique())
            machine_default = st.session_state.get("Machine", machine_options)
            machine_filter = col1.multiselect(
                "Machine",
                options=machine_options,
                default=[m for m in machine_default if m in machine_options],
                key="Machine"
            )
            df_view = df_view[df_view["Machine"].isin(machine_filter)]

            # ---------------- CHAMBER FILTER ----------------
            chamber_options = sorted(df_view_orig["Chamber"].dropna().unique())
            chamber_default = st.session_state.get("Chamber", chamber_options)
            chamber_filter = col2.multiselect(
                "Chamber",
                options=chamber_options,
                default=[c for c in chamber_default if c in chamber_options],
                key="Chamber"
            )
            df_view = df_view[df_view["Chamber"].isin(chamber_filter)]

            # ---------------- HOLE FILTER ----------------
            hole_options = sorted(df_view_orig["Hole"].dropna().unique())
            hole_default = st.session_state.get("Hole", hole_options)
            hole_filter = col3.multiselect(
                "Hole",
                options=hole_options,
                default=[h for h in hole_default if h in hole_options],
                key="Hole"
            )
            df_view = df_view[df_view["Hole"].isin(hole_filter)]

            # ---------------- FEATURE FILTER ----------------
            feature_options = sorted(df_view_orig["Feature"].dropna().unique())
            feature_default = [f for f in st.session_state.get("Feature", feature_options) if f in feature_options]
            feature_filter = col4.multiselect(
                "Feature",
                options=feature_options,
                default=feature_default if feature_default else [],
                key="Feature"
            )
            df_view = df_view[df_view["Feature"].isin(feature_filter)]

            # ---------------- PIECE ID FILTER ----------------
            piece_input = col5.text_input(
                "Piece ID (contains)",
                value=st.session_state.get("Piece_ID_filter", ""),
                key="Piece_ID_filter"
            )
            if piece_input.strip():
                df_view = df_view[df_view["Piece ID"].astype(str).str.contains(piece_input.strip(), case=False)]

            # ---------------- STATUS FILTER ----------------
            if "Status" in df_view.columns:
                df_view["Status"] = df_view["Status"].astype(str).str.strip().str.upper()
                status_options = ["PASS", "FAIL"]
                status_filter = col6.selectbox(
                    "Status",
                    options=["All"] + status_options,
                    index=0,
                    key="Status"
                )
                if status_filter != "All":
                    df_view = df_view[df_view["Status"] == status_filter.upper()]

            # ---------------- DATE RANGE FILTER ----------------
            if "Measured Date" in df_view.columns:
                valid_dates = df_view["Measured Date"].dropna()
                min_date = valid_dates.min().date() if not valid_dates.empty else datetime.now().date()
                max_date = valid_dates.max().date() if not valid_dates.empty else datetime.now().date()
                start_date, end_date = st.date_input(
                    "Measured Date Range",
                    value=(min_date, max_date)
                )
                df_view = df_view[
                    (df_view["Measured Date"].dt.date >= start_date) &
                    (df_view["Measured Date"].dt.date <= end_date)
                ]

        st.success(f"Filtered results: **{len(df_view)} rows**")

    # ===================== DISPLAY TABLE =====================
    import urllib.parse
    import base64

    if df_view.empty:
        st.info("No matching records found. Adjust filter or select another part.")
    else:
        df_display = df_view.copy()
        df_display.index = df_display.index + 1  # make 1-based like Excel

        # ---------- Build HTML Table ----------
        html = """
        <style>
            table.custom-table {
                border-collapse: collapse;
                width: 100%;
                font-family: Arial, sans-serif;
                font-size: 14px;
            }
            table.custom-table th, table.custom-table td {
                border: 1px solid #ccc;
                padding: 6px 8px;
                text-align: center;
            }
            table.custom-table th {
                background-color: #2c3e50;
                color: white;
                font-weight: bold;
            }
            .pass { background-color: #27ae60 !important; color: white !important; font-weight: bold; }
            .fail { background-color: #e74c3c !important; color: white !important; font-weight: bold; }
            tr.separator td { border-top: 4px solid black !important; }
            .image-cell img {
                max-width: 100px;
                max-height: 100px;
                display: block;
                margin: auto;
                cursor: pointer;
            }
            .value-cell { background-color: #f9f9d9; }
            .lsl-cell { background-color: #e8f4f8; }
            .usl-cell { background-color: #e8f8e8; }
        </style>
        <table class="custom-table">
            <thead><tr>
        """

        html += "<th>Row</th>" + "".join(f"<th>{col}</th>" for col in df_display.columns) + "</tr></thead><tbody>"

        for i in range(len(df_display)):
            row = df_display.iloc[i]
            row_class = ""

            if i > 0 and "Piece ID" in df_display.columns:
                prev_id = df_display.iloc[i - 1]["Piece ID"]
                curr_id = row["Piece ID"]
                if str(curr_id).strip() != str(prev_id).strip():
                    row_class = "separator"

            html += f"<tr class='{row_class}'><td><b>{i+1}</b></td>"

            for col in df_display.columns:
                val = row[col]
                val = "" if pd.isna(val) else val
                cell_class = ""

                if col.lower() == "status":
                    cell_class = "pass" if str(val).lower() == "pass" else "fail"

                if col == "Value": cell_class += " value-cell"
                elif col == "LSL": cell_class += " lsl-cell"
                elif col == "USL": cell_class += " usl-cell"

                if col == "Image Path" and val != "":
                    img_path = val.strip()
                    if os.path.exists(img_path):
                        with open(img_path, "rb") as f:
                            data = f.read()
                            encoded = base64.b64encode(data).decode()
                        html += f"<td class='{cell_class}'><img src='data:image/jpeg;base64,{encoded}' class='image-cell'/></td>"
                    else:
                        html += f"<td class='{cell_class}'></td>"
                else:
                    html += f"<td class='{cell_class}'>{val}</td>"

            html += "</tr>"

        html += "</tbody></table>"
        st.markdown(html, unsafe_allow_html=True)

    # ===================== DELETE ROWS =====================
    st.markdown("### 🗑️ Delete Rows")
    st.info("Enter row numbers exactly as shown in the table (1-based). Example: `1,2` or `3-5,7`")
    delete_input = st.text_input("Rows to delete (e.g. 1,2 or 1-5,7)", key="delete_input_view")

    if st.button("Delete Selected"):
        raw = str(delete_input).strip()
        if not raw:
            st.warning("Enter rows or ranges to delete.")
        else:
            try:
                # Parse user input
                rows_to_delete = []
                parts = [r.strip() for r in raw.split(",") if r.strip()]
                for p in parts:
                    if "-" in p:
                        start, end = p.split("-")
                        rows_to_delete.extend(range(int(start), int(end) + 1))
                    else:
                        rows_to_delete.append(int(p))

                rows_to_delete = sorted(set(rows_to_delete))
                # Map filtered indices to original Excel indices
                rows_to_delete_orig = df_view["_orig_index"].iloc[[r-1 for r in rows_to_delete if 1 <= r <= len(df_view)]].tolist()

                if not rows_to_delete_orig:
                    st.warning("⚠️ No valid rows to delete.")
                else:
                    df_upd = df_view_orig.drop(index=rows_to_delete_orig).reset_index(drop=True)

                    # Save back to Excel
                    try:
                        with pd.ExcelWriter(EXCEL, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
                            df_upd.to_excel(writer, sheet_name=sheet_name, index=False)

                        st.success(f"✅ Deleted rows: {rows_to_delete}")
                        st.toast(f"Rows {rows_to_delete} deleted successfully 🗑️", icon="🗑️")
                        import time
                        time.sleep(1.5)
                        st.rerun()
                    except PermissionError:
                        st.error("❌ Excel file is open. Please **close Excel first** and try again.")
                    except Exception as e:
                        st.error(f"❌ Failed to save changes: {e}")

            except Exception as e:
                st.error(f"❌ Failed to delete rows: {e}")

    # ===================== DELETE IMAGE FOR SPECIFIC ROW =====================
    st.markdown("### 🖼️ Delete Image for Selected Hole")
    st.info("Enter row numbers exactly as shown in the table (1-based). Example: `1,2` or `3-5,7`")
    delete_image_input = st.text_input("Rows to delete image for the selected hole (e.g. 1,2 or 1-5,7)", key="delete_image_input_view")

    if st.button("Delete Image for Selected Holes"):
        raw = str(delete_image_input).strip()
        if not raw:
            st.warning("Enter rows or ranges to delete the images.")
        else:
            try:
                rows_to_delete_image = []
                parts = [r.strip() for r in raw.split(",") if r.strip()]
                for p in parts:
                    if "-" in p:
                        start, end = p.split("-")
                        rows_to_delete_image.extend(range(int(start), int(end) + 1))
                    else:
                        rows_to_delete_image.append(int(p))

                rows_to_delete_image = sorted(set(rows_to_delete_image))
                rows_to_delete_image_orig = df_view["_orig_index"].iloc[[r-1 for r in rows_to_delete_image if 1 <= r <= len(df_view)]].tolist()

                if not rows_to_delete_image_orig:
                    st.warning("⚠️ No valid rows to delete image for.")
                else:
                    df_upd = df_view_orig.copy()
                    for idx_orig in rows_to_delete_image_orig:
                        image_path = df_upd.at[idx_orig, 'Image Path']
                        if image_path and os.path.exists(image_path):
                            try:
                                os.remove(image_path)
                            except Exception as e:
                                st.error(f"❌ Failed to delete image: {e}")
                        df_upd.at[idx_orig, 'Image Path'] = None

                    # Save updated Excel
                    try:
                        with pd.ExcelWriter(EXCEL, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
                            df_upd.to_excel(writer, sheet_name=sheet_name, index=False)
                        st.success("✅ Images deleted successfully.")
                        st.toast("Selected images deleted successfully 🖼️", icon="🖼️")
                        import time
                        time.sleep(1.5)
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Failed to save changes: {e}")
            except Exception as e:
                st.error(f"❌ Failed to delete images: {e}")

    # ===================== EDIT MEASUREMENT DATA =====================
    st.markdown("### ✏️ Edit Measurement Data ")
    if df_view.empty:
        st.info("No data to edit.")
    else:
        total_rows = len(df_view)
        st.info(f"Enter the **row number** as shown in the table (1 to {total_rows}).")
        edit_row_display = st.number_input(
            "Enter Row Number to Edit (Press ENTER to edit the row number)",
            min_value=1,
            max_value=total_rows,
            value=1,
            step=1
        )
        edit_row = edit_row_display - 1
        selected_row = df_view.iloc[edit_row]
        st.markdown(f"**Editing Row {edit_row_display}** — Piece ID: `{selected_row.get('Piece ID', 'N/A')}`")

        editable_cols = [
            "Measured Date", "Batch Cleaning", "Machine", "Part Type", "Chamber",
            "Piece ID", "Hole", "Feature", "Value",
            "Part In/Out", "Notes", "Image Path"
        ]
        editable_cols = [c for c in editable_cols if c in df_view.columns]

        with st.form("form_edit_row_excel_like"):
            new_entries = {}
            for col in editable_cols:
                val = selected_row[col]
                if col == "Measured Date":
                    try:
                        date_val = pd.to_datetime(val).date() if pd.notna(val) else datetime.now().date()
                    except:
                        date_val = datetime.now().date()
                    new_entries[col] = st.date_input("Measured Date", value=date_val, key=f"edit_{col}")
                    
                elif col == "Batch Cleaning":
                    # Ensure value is float to match min_value and step type
                    batch_val = 0.0 if val in [None, ""] else float(val)
                    new_entries[col] = st.number_input(
                        "Batch Cleaning (optional)",
                        value=batch_val,
                        min_value=0.0,
                        max_value=None,  # explicitly allow None
                        step=1.0,
                        format="%.0f",
                        key=f"edit_{col}"
                    )

                else:
                    new_entries[col] = st.text_input(col, value="" if pd.isna(val) else str(val), key=f"edit_{col}")

            part_inout_val = selected_row.get("Part In/Out", "IN")
            edit_part_inout = st.selectbox("Part In/Out", options=["IN", "OUT"], index=0 if str(part_inout_val).upper() == "IN" else 1)

            hole_type = selected_row.get('Hole', 'N/A')
            hole_image = st.file_uploader(f"Upload Image for Hole {hole_type}", type=["png", "jpg", "jpeg"], key=f"hole_{hole_type}")
            if hole_image:
                feature_type = selected_row.get("Feature", "Unknown")
                image_filename = f"{selected_row['Hole']}_{feature_type}.jpg"
                image_path = os.path.join("uploaded_images", image_filename)
                with open(image_path, "wb") as img_file:
                    img_file.write(hole_image.getbuffer())
                new_entries["Image Path"] = image_path
                st.markdown(f'<a href="file:///{os.path.abspath(image_path)}" target="_blank">'
                            f'<img src="file:///{os.path.abspath(image_path)}" width="400" style="border-radius: 10px; cursor: pointer;" /></a>',
                            unsafe_allow_html=True)

            save_clicked = st.form_submit_button("💾 Save Changes")

        if save_clicked:
            # Map edit to original Excel index
            orig_idx = selected_row["_orig_index"]
            df_upd = df_view_orig.copy()
            for col, val in new_entries.items():
                if col == "Measured Date":
                    df_upd.at[orig_idx, col] = val.strftime("%Y-%m-%d")
                elif col == "Batch Cleaning":
                    df_upd.at[orig_idx, col] = val if val != 0 else ""
                elif col == "Value":
                    try:
                        df_upd.at[orig_idx, col] = float(val)
                    except:
                        df_upd.at[orig_idx, col] = val
                else:
                    df_upd.at[orig_idx, col] = val

            df_upd.at[orig_idx, "Part In/Out"] = edit_part_inout

            # Recalculate Status
            try:
                val = df_upd.at[orig_idx, "Value"]
                lsl = df_upd.at[orig_idx, "LSL"]
                usl = df_upd.at[orig_idx, "USL"]
                if pd.notna(val) and pd.notna(lsl) and pd.notna(usl):
                    df_upd.at[orig_idx, "Status"] = "Fail" if val < lsl or val > usl else "Pass"
                else:
                    df_upd.at[orig_idx, "Status"] = "N/A"
            except:
                df_upd.at[orig_idx, "Status"] = "N/A"

            try:
                with pd.ExcelWriter(EXCEL, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
                    df_upd.to_excel(writer, sheet_name=sheet_name, index=False)
                st.success(f"✅ Row {edit_row_display} updated successfully in Excel and table.")
                st.toast(f"Row {edit_row_display} updated successfully ✅", icon="✅")
                import time
                time.sleep(1.5)
                st.rerun()
            except Exception as e:
                st.error(f"❌ Failed to save changes: {e}")

                    
# ------------------ TAB 3: View Spec ------------------
with tabs[3]:
    st.subheader("Specs — Limits & Tolerances")

    # ⚠️ Universal Auto-Hiding Excel Warning Banner
    warning_html = """
    <div id="excel-warning" style="
        background-color:#ffe6e6;
        padding:15px;
        border-radius:10px;
        box-shadow:0px 4px 10px rgba(255,0,0,0.3);
        text-align:center;
        font-weight:bold;
        color:#990000;
        font-size:18px;
        margin-bottom:15px;
        animation: fadeOut 6s forwards;">
        🚨 <u>IMPORTANT:</u> PLEASE CLOSE THE EXCEL FILE FIRST BEFORE ADDING, EDITING, OR DELETING DATA!
    </div>

    <style>
    @keyframes fadeOut {
      0%   { opacity: 1; }
      80%  { opacity: 1; }
      100% { opacity: 0; display: none; }
    }
    </style>
    """

    def show_excel_warning():
        st.markdown(warning_html, unsafe_allow_html=True)

    # Part filter for Specs
    spec_part = st.selectbox("Part Type", ["All", "Mixing Block", "Gas/Water Block"], key="spec_part")
    df_specs = get_specs_df(spec_part)

    if df_specs is None or df_specs.empty:
        st.info("Specs not available yet.")
    else:
        # Show the specs table
        st.dataframe(df_specs, use_container_width=True)

        # Export to vendor CSV
        col_left, col_right = st.columns([2, 1])
        with col_left:
            filename_input = st.text_input("Export filename", value="specs.csv", key="spec_export_name")
        with col_right:
            if st.button("Export specs"):
                ok, result = export_specs_for_vendor(
                    filename=filename_input,
                    part_filter=(spec_part if spec_part != "All" else None)
                )
                if ok:
                    st.success(f"Specs exported: {result}")
                    try:
                        with open(result, "rb") as fh:
                            st.download_button("📥 Download Exported CSV", fh.read(), file_name=os.path.basename(result))
                    except Exception:
                        st.info("Export saved; file may be on server filesystem.")
                else:
                    st.error(f"Export failed: {result}")

# ------------------ TAB 4: Reference Photo ------------------
with tabs[4]:
    show_reference_photos()

# ------------------ TAB 5: Workbook ------------------
with tabs[5]:
    st.subheader("Excel Files")

    # --- Excel file paths ---
    TREND_EXCEL = os.path.join(os.getcwd(), "trendchart.xlsx")  # Ensure path is relative to current folder

    excel_options = {
        "Main Data Workbook": EXCEL,
        "Trend Charts Workbook": TREND_EXCEL
    }

    selected_file_label = st.selectbox("Select Excel File to Open", list(excel_options.keys()))
    selected_file_path = excel_options[selected_file_label]

    # --- Open / Download Columns ---
    col_open, col_dl = st.columns([1, 2])

    with col_open:
        if st.button(f"Open '{selected_file_label}' in Excel"):
            if os.path.exists(selected_file_path):
                try:
                    os.startfile(selected_file_path)  # Windows only
                    st.success(f"✅ '{selected_file_label}' opened successfully!")
                except Exception as e:
                    st.error(f"Could not open '{selected_file_label}': {e}")
            else:
                st.warning(f"File '{selected_file_label}' does not exist. Please generate it first.")

    with col_dl:
        try:
            if os.path.exists(selected_file_path):
                with open(selected_file_path, "rb") as fh:
                    data = fh.read()
                st.download_button(
                    f"📥 Download '{selected_file_label}'",
                    data,
                    file_name=os.path.basename(selected_file_path)
                )
            else:
                st.warning(f"File '{selected_file_label}' not found. Cannot download.")
        except Exception as e:
            st.error(f"File '{selected_file_label}' not ready: {e}")

# Footer
st.markdown("<hr>", unsafe_allow_html=True)
st.markdown("<small style='color:#999'>© 2025 — SA PM Logger • Test6</small>", unsafe_allow_html=True)

theme_css = """
<style>
div[data-testid="stAppViewContainer"] > header h1,
div[data-testid="stAppViewContainer"] > header h2,
div[data-testid="stHeader"] h1,
div[data-testid="stHeader"] h2 {
    color: #0033cc !important; /* ✅ Dark Blue */
}
</style>
"""
st.markdown(theme_css, unsafe_allow_html=True)

# ------------------ TAB 6: Professional Summary Dashboard ------------------
with tabs[6]:

    import pandas as pd
    import numpy as np
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    import plotly.graph_objects as go

    st.markdown("""
        <div style="padding:15px; background-color:#e6f2ff; border-radius:10px;">
            <h2 style="text-align:center; color:#003366;">🏭 FUJI ELECTRIC – SA-CVD PMSA SPEC CONTROL SYSTEM</h2>
            <h3 style="text-align:center; color:#003366;">📘 Measurement Summary Report</h3>
        </div>
    """, unsafe_allow_html=True)

    st.write("")
    st.write("This dashboard provides a overview of Mixing Block and Gas-Water Block measurements with automatic spec checking, KPI metrics, and Excel export.")

    # ------------------ Helper Functions ------------------
    def classify_hole(hole_name):
        hole_name = str(hole_name).upper()
        if hole_name in ["H1", "H2", "H3", "H4"]:
            return "Inner"
        elif hole_name in ["H5", "H6", "H7", "H8", "H9", "H10"]:
            return "Outer"
        return "Unknown"

    def highlight_spec(val):
        if val == "No":
            return "background-color:#f9c0c0"
        return "background-color:#d4f4dd"

    # ------------------ Generate Report Button ------------------
    if st.button("🟢 Generate Summary Report"):

        with st.spinner("⏳ Generating your summary report..."):

            try:
                # Load Excel Data
                df_mb = pd.read_excel("test6.xlsx", sheet_name="Mixing Block Data")
                df_gwb = pd.read_excel("test6.xlsx", sheet_name="Gas-Water Block Data")

                for df in [df_mb, df_gwb]:
                    df.fillna("", inplace=True)
                    df["Value"] = pd.to_numeric(df["Value"], errors="coerce")
                    df["Hole Type"] = df["Hole"].apply(classify_hole)
                    df["In Spec"] = np.where(
                        (df["Value"] >= df["LSL"]) & (df["Value"] <= df["USL"]),
                        "Yes", "No"
                    )

                # ------------------ KPI Metrics ------------------
                total_measurements = len(df_mb) + len(df_gwb)
                total_oos = (df_mb["In Spec"] == "No").sum() + (df_gwb["In Spec"] == "No").sum()
                in_spec_rate = (1 - total_oos / total_measurements) * 100

                col1, col2, col3 = st.columns(3)
                col1.metric("📏 Total Measurements", total_measurements)
                col2.metric("🟢 In-Spec Rate", f"{in_spec_rate:.2f}%")
                col3.metric("🔴 Out-of-Spec Count", total_oos)

                st.markdown("---")

                display_cols = ["Hole Type", "Hole", "Piece ID", "Part Type",
                                "Value", "LSL", "USL", "In Spec"]

                # ------------------ MIXING BLOCK ------------------
                st.markdown("## 🟦 Mixing Block Summary")
                df_mb_show = df_mb[display_cols].copy()
                df_mb_show.index = np.arange(1, len(df_mb_show)+1)
                st.dataframe(df_mb_show.style.applymap(highlight_spec, subset=["In Spec"]))

                st.markdown("### 🔴 Out of Spec (Mixing Block)")
                df_mb_oos = df_mb[df_mb["In Spec"] == "No"][display_cols].copy()
                df_mb_oos.index = df_mb_show.index[df_mb["In Spec"] == "No"]
                if df_mb_oos.empty:
                    st.success("All Mixing Block measurements are within specification.")
                else:
                    st.dataframe(df_mb_oos.style.applymap(highlight_spec, subset=["In Spec"]))

                st.markdown("---")

                # ------------------ GAS WATER BLOCK ------------------
                st.markdown("## 🟩 Gas-Water Block Summary")
                df_gwb_show = df_gwb[display_cols].copy()
                df_gwb_show.index = np.arange(1, len(df_gwb_show)+1)
                st.dataframe(df_gwb_show.style.applymap(highlight_spec, subset=["In Spec"]))

                st.markdown("### 🔴 Out of Spec (Gas-Water Block)")
                df_gwb_oos = df_gwb[df_gwb["In Spec"] == "No"][display_cols].copy()
                df_gwb_oos.index = df_gwb_show.index[df_gwb["In Spec"] == "No"]
                if df_gwb_oos.empty:
                    st.success("All Gas-Water Block measurements are within specification.")
                else:
                    st.dataframe(df_gwb_oos.style.applymap(highlight_spec, subset=["In Spec"]))

                st.markdown("---")

                # ------------------ In-Spec vs Out-of-Spec Chart ------------------
                fig = go.Figure()
                fig.add_trace(go.Bar(
                    x=["Mixing Block", "Gas-Water Block"],
                    y=[len(df_mb[df_mb["In Spec"]=="Yes"]), len(df_gwb[df_gwb["In Spec"]=="Yes"])],
                    name="In Spec", marker_color="#2ca02c"
                ))
                fig.add_trace(go.Bar(
                    x=["Mixing Block", "Gas-Water Block"],
                    y=[len(df_mb[df_mb["In Spec"]=="No"]), len(df_gwb[df_gwb["In Spec"]=="No"])],
                    name="Out of Spec", marker_color="#d62728"
                ))

                fig.update_layout(
                    barmode='stack',
                    title=dict(
                        text="📊 In-Spec vs Out-of-Spec Measurements",
                        font=dict(size=18, color="#003366")
                    ),
                    xaxis=dict(
                        title=dict(
                            text="Block Type",
                            font=dict(size=14, color="#003366")
                        )
                    ),
                    yaxis=dict(
                        title=dict(
                            text="Count",
                            font=dict(size=14, color="#003366")
                        )
                    ),
                    legend=dict(title="Status", font=dict(size=12)),
                    template="plotly_white",
                    margin=dict(l=60, r=40, t=80, b=40),
                    height=400
                )
                st.plotly_chart(fig, use_container_width=True)

                # ------------------ CREATE PROFESSIONAL EXCEL ------------------
                def write_sheet(ws, df, title):
                    # Add Index column starting from 1
                    df_insert = df.copy()
                    df_insert.insert(0, "Index", np.arange(1, len(df_insert)+1))

                    # Merge top row for title
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_insert.columns))
                    cell = ws.cell(row=1, column=1, value=title)
                    cell.font = Font(size=14, bold=True, color="003366")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(start_color="cce6ff", fill_type="solid")

                    # Leave one blank row
                    ws.append([])

                    # Header styling
                    header_fill = PatternFill(start_color="ADD8E6", fill_type="solid")
                    header_font = Font(bold=True, color="003366")
                    thin_border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))

                    for c, col in enumerate(df_insert.columns, 1):
                        cell = ws.cell(row=3, column=c, value=col)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                        cell.border = thin_border

                    # Write data rows
                    for r, row in enumerate(df_insert.itertuples(index=False), 4):
                        for c, val in enumerate(row, 1):
                            cell = ws.cell(row=r, column=c, value=val)
                            cell.alignment = Alignment(horizontal="center")
                            cell.border = thin_border
                            if df_insert.columns[c-1] == "In Spec":
                                if val == "No":
                                    cell.fill = PatternFill(start_color="F9C0C0", fill_type="solid")
                                else:
                                    cell.fill = PatternFill(start_color="D4F4DD", fill_type="solid")


                buffer = BytesIO()
                wb = Workbook()

                ws1 = wb.active
                ws1.title = "Mixing Block - All"
                write_sheet(ws1, df_mb_show, "Mixing Block - All Measurements")

                ws2 = wb.create_sheet("Mixing Block - OOS")
                write_sheet(ws2, df_mb_oos, "Mixing Block - Out of Spec Measurements")

                ws3 = wb.create_sheet("Gas Water Block - All")
                write_sheet(ws3, df_gwb_show, "Gas-Water Block - All Measurements")

                ws4 = wb.create_sheet("Gas Water Block - OOS")
                write_sheet(ws4, df_gwb_oos, "Gas-Water Block - Out of Spec Measurements")

                wb.save(buffer)
                buffer.seek(0)

                st.success("✅ Summary report generated successfully!")
                st.download_button(
                    label="📥 Download Summary Excel",
                    data=buffer,
                    file_name="SA_CVD_Summary.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            except Exception as e:
                st.error(f"⚠️ Error: {e}")
